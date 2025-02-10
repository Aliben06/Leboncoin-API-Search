import pandas as pd
import numpy as np
from datetime import datetime
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor
import multiprocessing
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

class AnalyseurVoitures:
    def __init__(self, fichier_csv):
        """Initialise l'analyseur avec le fichier CSV"""
        print("Chargement des données...")
        self.df = self._charger_donnees(fichier_csv)
        self.prix_moyens_marche = {}
        print(f"Données chargées : {len(self.df)} véhicules trouvés")
        
    def _charger_donnees(self, fichier_csv):
        """Charge et nettoie les données avec indication de progression"""
        try:
            df = pd.read_csv(fichier_csv, encoding='utf-8')
        except UnicodeDecodeError:
            df = pd.read_csv(fichier_csv, encoding='latin1')
        
        print("Nettoyage des données...")
        for colonne in ['Prix', 'Kilométrage', 'Puissance din', 'Année']:
            if colonne in df.columns:
                print(f"Traitement de la colonne {colonne}...")
                if colonne in ['Prix', 'Kilométrage', 'Puissance din']:
                    df[colonne] = pd.to_numeric(df[colonne].astype(str).str.replace('[^0-9.]', '', regex=True), errors='coerce')
                else:
                    df[colonne] = pd.to_numeric(df[colonne], errors='coerce')

        # Supprimer les valeurs invalides
        print("Validation des données...")
        df = df[df['Prix'] > 0]
        df = df[df['Kilométrage'] > 0]
        df = df[df['Année'] > 0]

        return df
    
    def analyser_tendances_marche(self):
        """Analyse les tendances du marché pour chaque modèle"""
        tendances = {}
        
        for (marque, modele), groupe in self.df.groupby(['Marque', 'Modele']):
            if len(groupe) < 3:
                continue
                
            tendances[f"{marque} {modele}"] = {
                'prix_moyen': groupe['Prix'].mean(),
                'prix_median': groupe['Prix'].median(),
                'ecart_type_prix': groupe['Prix'].std(),
                'km_moyen': groupe['Kilométrage'].mean(),
                'age_moyen': datetime.now().year - groupe['Année'].mean(),
                'nombre_annonces': len(groupe),
                'prix_min': groupe['Prix'].min(),
                'prix_max': groupe['Prix'].max(),
            }
            
        return tendances

    def calculer_score_value(self, vehicule):
        """Calcule un score de rapport qualité-prix"""
        groupe = self.df[
            (self.df['Marque'] == vehicule['Marque']) &
            (self.df['Modele'] == vehicule['Modele']) &
            (abs(self.df['Année'] - vehicule['Année']) <= 2)
        ]
        
        if len(groupe) < 2:
            return 0
        
        # Prix moyen au kilomètre pour ce modèle
        prix_km_moyen = groupe['Prix'].mean() / groupe['Kilométrage'].mean()
        prix_km_vehicule = vehicule['Prix'] / vehicule['Kilométrage']
        
        # Score basé sur l'écart par rapport à la moyenne
        score = 1 - (prix_km_vehicule / prix_km_moyen)
        
        return score

    def calculer_scores_par_batch(self, vehicules_batch):
        """Calcule les scores pour un batch de véhicules"""
        resultats = []
        for _, vehicule in vehicules_batch.iterrows():
            score = self.calculer_score_value(vehicule)
            if score > 0.25:  # Seuil ajusté à 0.25
                resultats.append({
                    'ID': vehicule['ID'],
                    'Marque': vehicule['Marque'],
                    'Modele': vehicule['Modele'],
                    'Prix': vehicule['Prix'],
                    'Score_value': score,
                    'Année': vehicule['Année'],
                    'Kilométrage': vehicule['Kilométrage'],
                    'URL': vehicule['URL'],
                    'Type': 'Bon rapport qualité-prix'
                })
        return resultats

    def detecter_anomalies_prix(self):
        """Détecte les véhicules dont le prix est anormalement bas"""
        print("\nAnalyse des anomalies de prix...")
        resultats = []
        groupes = list(self.df.groupby(['Marque', 'Modele']))
        seuil_ecart_type = 10_000  # Seuil pour exclure les groupes très variables

        with tqdm(total=len(groupes), desc="Analyse des groupes") as pbar:
            for (marque, modele), groupe in groupes:
                taille_groupe = len(groupe)
                if taille_groupe < 3 or groupe['Prix'].std() > seuil_ecart_type:
                    pbar.update(1)
                    continue

                # Calculs pour ce groupe
                prix_moyen = groupe['Prix'].mean()
                km_impact = -0.1 * (groupe['Kilométrage'] - groupe['Kilométrage'].mean()) / 10000
                annee_impact = 0.05 * (groupe['Année'] - groupe['Année'].mean())
                prix_predit = prix_moyen * (1 + km_impact + annee_impact)

                # Identifier les bonnes affaires
                for idx, row in groupe.iterrows():
                    prix_predit_vehicule = prix_predit.loc[idx]
                    if row['Prix'] < 0.75 * prix_predit_vehicule:  # Seuil ajusté à 75%
                        resultats.append({
                            'ID': row['ID'],
                            'Marque': row['Marque'],
                            'Modele': row['Modele'],
                            'Prix': row['Prix'],
                            'Prix_predit': prix_predit_vehicule,
                            'Économie': prix_predit_vehicule - row['Prix'],
                            'Pourcentage_économie': ((prix_predit_vehicule - row['Prix']) / prix_predit_vehicule) * 100,
                            'Année': row['Année'],
                            'Kilométrage': row['Kilométrage'],
                            'URL': row['URL'],
                            'Nombre_comparables': taille_groupe  # Taille du groupe
                        })
                pbar.update(1)

        print(f"Nombre d'anomalies trouvées : {len(resultats)}")
        return resultats

    def trouver_bonnes_affaires(self, criteres_personnalises=None):
        """Trouve les bonnes affaires avec traitement parallèle"""
        print("\nRecherche des bonnes affaires...")
        bonnes_affaires = []
        
        # Analyse des prix anormalement bas
        anomalies = self.detecter_anomalies_prix()
        
        # Calcul parallèle des scores value
        print("\nCalcul des scores qualité-prix...")
        n_cpu = multiprocessing.cpu_count()
        batch_size = max(1000, len(self.df) // (n_cpu * 4))
        batches = [self.df[i:i + batch_size] for i in range(0, len(self.df), batch_size)]
        
        with ThreadPoolExecutor(max_workers=n_cpu) as executor:
            futures = []
            for batch in batches:
                futures.append(executor.submit(self.calculer_scores_par_batch, batch))
            
            with tqdm(total=len(futures), desc="Traitement des batches") as pbar:
                for future in futures:
                    bonnes_affaires.extend(future.result())
                    pbar.update(1)
        
        # Ajouter les anomalies de prix
        bonnes_affaires.extend([{**anomalie, 'Type': 'Prix anormalement bas'} for anomalie in anomalies])
        
        # Appliquer les critères personnalisés
        if criteres_personnalises:
            print("\nApplication des critères personnalisés...")
            bonnes_affaires = [ba for ba in tqdm(bonnes_affaires) if self._verifier_criteres(ba, criteres_personnalises)]
        
        # Tri final
        print("\nTri des résultats...")
        bonnes_affaires.sort(key=lambda x: x.get('Pourcentage_économie', x.get('Score_value', 0)), reverse=True)
        
        return bonnes_affaires

    def exporter_resultats(self, bonnes_affaires, tendances, nom_fichier='resultats_analyse.xlsx'):
            """Exporte les résultats dans un fichier Excel formaté"""
            print(f"\nExportation des résultats vers {nom_fichier}...")
            
            # Créer un writer Excel avec openpyxl
            with pd.ExcelWriter(nom_fichier, engine='openpyxl') as writer:
                # 1. Export des bonnes affaires
                df_bonnes_affaires = pd.DataFrame(bonnes_affaires)
                if not df_bonnes_affaires.empty:
                    # Réorganiser les colonnes
                    colonnes = ['Type', 'Marque', 'Modele', 'Prix', 'Prix_predit', 'Économie', 
                            'Pourcentage_économie', 'Score_value', 'Année', 'Kilométrage', 'URL']
                    df_bonnes_affaires = df_bonnes_affaires.reindex(columns=[col for col in colonnes if col in df_bonnes_affaires.columns])
                    
                    # Formater les nombres
                    for col in ['Prix', 'Prix_predit', 'Économie']:
                        if col in df_bonnes_affaires.columns:
                            df_bonnes_affaires[col] = df_bonnes_affaires[col].round(2)
                    
                    df_bonnes_affaires.to_excel(writer, sheet_name='Bonnes Affaires', index=False)
                    
                    # Formater la feuille
                    sheet = writer.sheets['Bonnes Affaires']
                    for idx, col in enumerate(df_bonnes_affaires.columns, 1):
                        cell = sheet.cell(row=1, column=idx)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                        sheet.column_dimensions[get_column_letter(idx)].width = 15
                    
                # 2. Export des tendances du marché
                df_tendances = pd.DataFrame.from_dict(tendances, orient='index')
                df_tendances.to_excel(writer, sheet_name='Tendances Marché')
                
                # 3. Export des statistiques globales
                stats_globales = {
                    'Métrique': [
                        'Nombre total de véhicules',
                        'Nombre de bonnes affaires',
                        'Prix moyen',
                        'Prix médian',
                        'Kilométrage moyen',
                        'Année moyenne'
                    ],
                    'Valeur': [
                        len(self.df),
                        len(bonnes_affaires),
                        self.df['Prix'].mean(),
                        self.df['Prix'].median(),
                        self.df['Kilométrage'].mean(),
                        self.df['Année'].mean()
                    ]
                }
                df_stats = pd.DataFrame(stats_globales)
                df_stats.to_excel(writer, sheet_name='Statistiques Globales', index=False)
                
                # 4. Export des top modèles
                top_modeles = (self.df.groupby(['Marque', 'Modele'])
                            .agg({
                                'Prix': ['count', 'mean', 'min', 'max'],
                                'Kilométrage': 'mean',
                                'Année': 'mean'
                            })
                            .reset_index())
                top_modeles.columns = ['Marque', 'Modele', 'Nombre_annonces', 'Prix_moyen', 
                                    'Prix_min', 'Prix_max', 'Km_moyen', 'Année_moyenne']
                top_modeles = top_modeles.sort_values('Nombre_annonces', ascending=False)
                top_modeles.to_excel(writer, sheet_name='Top Modèles', index=False)
                
                # 5. Export des anomalies de prix
                anomalies = [item for item in bonnes_affaires if item['Type'] == 'Prix anormalement bas']
                if anomalies:
                    df_anomalies = pd.DataFrame(anomalies)
                    df_anomalies.to_excel(writer, sheet_name='Anomalies Prix', index=False)

            print(f"Résultats exportés avec succès dans {nom_fichier}")
            return nom_fichier

    def generer_rapport_complet(self):
        """Génère et exporte un rapport complet d'analyse"""
        print("\nGénération du rapport complet...")
        
        # Analyser les tendances
        print("Analyse des tendances du marché...")
        tendances = self.analyser_tendances_marche()
        
        # Trouver les bonnes affaires
        print("Recherche des bonnes affaires...")
        bonnes_affaires = self.trouver_bonnes_affaires()
        
        # Exporter les résultats
        nom_fichier = f'analyse_voitures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        fichier_resultat = self.exporter_resultats(bonnes_affaires, tendances, nom_fichier)
        
        print(f"\nAnalyse terminée! Les résultats ont été enregistrés dans {fichier_resultat}")
        
        # Afficher un résumé
        print(f"\nRésumé de l'analyse:")
        print(f"- Nombre total de véhicules analysés: {len(self.df)}")
        print(f"- Nombre de bonnes affaires trouvées: {len(bonnes_affaires)}")
        print(f"- Nombre de modèles différents: {len(tendances)}")
        
        return fichier_resultat


# Exemple d'utilisation
if __name__ == "__main__":
    try:
        print("Démarrage de l'analyse...")
        analyseur = AnalyseurVoitures('resume_2025-01-13__BretagneV2.csv')
        fichier_resultat = analyseur.generer_rapport_complet()
        print(f"\nVous pouvez maintenant ouvrir {fichier_resultat} pour voir les résultats détaillés.")
    except Exception as e:
        print(f"Une erreur est survenue: {str(e)}")